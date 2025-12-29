from pathlib import Path
import webbrowser
import time
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import subprocess
def update():
    # Descarga el Excel
    url = "https://vertivco-my.sharepoint.com/personal/julian_garcia1_vertivco_com/_layouts/15/download.aspx?UniqueId=cc4f8e90%2D9ecb%2D4f2d%2Da368%2D07d5b8d18a40"
    webbrowser.open(url)
    time.sleep(7)
    # Busca Excel descargado
    carpeta = Path(r"C:\Users\Miguel.Carranza\Downloads")
    archivo = carpeta / "Personal_Mty_2.0.xlsx"

    if archivo.exists() and archivo.is_file():
        print("Archivo encontrado:", archivo)
    else:
        print("No está (todavía).")

    # Copiar de Personal_MTY_2.0
    hoja = "HC 16-10-25_TeamsDraft"
    columnas = ["A", "B", "C", "J", "K", "L"]
    archivo_salida = r"C:\Inventario\data\roles_areas.xlsx"

    wb = load_workbook(archivo, data_only=True)
    ws = wb[hoja]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Filtrado"

    # Convertir letras a índices
    idxs = [column_index_from_string(c) for c in columnas]

    # Copiar todas las filas usadas (1..max_row)
    for fila in range(1, ws.max_row + 1):
        for j, col_idx in enumerate(idxs, start=1):
            out_ws.cell(row=fila, column=j).value = ws.cell(row=fila, column=col_idx).value

    out_wb.save(archivo_salida)
    print("✅ Listo:", archivo_salida)


    # roles_area
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    hoja = "Filtrado"
    col_loc = "D" 

    MAP = {
        "DH L2": "DH_400",
        "SKID": "SKID",
        "MC": "MC",
        "BLOWER": "Blower",
        "SHP": "SHIP",
        "COILS": "Coils_Shop",
        "XDU1350": "XDU_1350",
        "KITS": "Kits",
        "MTY 1": "Warehouse",
        "3PL": "MTY3",
        "DA / PLENUM": "MONDA",
        "DH L1" : "DH_400",
        "COMPRESOR": "CR",
        "XDU100": "XDU_100",
        "PIPING": "Piping",
        "MCV" : "MC",
        "PANELES": "EP",
        "PRE": "PRE_L1",
        "XDM": "XDM",
        "AFC": "AFC",
        "WIP": "WIP",
        "MRB": "MRB",
        "ALL": "ALL",
        "POLO": "IT",
        "SHEET METAL": "Sheet_Metal",
        "APU1": "APU1",
        "PLANTA" : "ALL",
        "FASTENAL": "Fastenal",
        "LAREDO": "Laredo",
        "APU2 APU3": "APU2 APU3",
        "MONBLW": "Blower",
        "MONCR": "CR",
        "MONDA": "MONDA",
        "MONEP": "EP",
        "MONPIPING": "Piping",
        "MONSM": "Sheet_Metal",
        "SHIP": "SHIP",
    }

    def norm(x):
        if x is None:
            return ""
        return " ".join(str(x).strip().split()).upper()  # quita dobles espacios y normaliza a MAYÚSC

    wb = load_workbook(archivo_salida)
    ws = wb[hoja] if hoja else wb.active

    ws.insert_cols(6)
    ws.cell(row=1, column=6).value = "SVG_ID"

    loc_col_idx = column_index_from_string(col_loc)

    # Rellenar SVG_ID fila por fila (desde la 2)
    for r in range(2, ws.max_row + 1):
        key = norm(ws.cell(row=r, column=loc_col_idx).value)
        ws.cell(row=r, column=6).value = MAP.get(key, "N/A")  # si no hay match, queda vacío

    wb.save(archivo_salida)
    print("✅ Listo: inserté SVG_ID en la columna F y apliqué equivalencias.")

    #Github
    REPO = Path(r"C:\Inventario")   # carpeta donde está el .git
    MENSAJE = "Auto update: roles_areas.xlsx"

    def run(cmd: list[str]) -> None:
        r = subprocess.run(cmd, cwd=REPO, text=True, capture_output=True)
        if r.returncode != 0:
            raise RuntimeError(f"Error ejecutando: {' '.join(cmd)}\n{r.stderr}")
        if r.stdout.strip():
            print(r.stdout.strip())

    # 1) (opcional) ver estado
    run(["git", "status"])

    # 2) add (todo) o solo un archivo
    # run(["git", "add", "."])
    run(["git", "add", "."])  # pon la ruta relativa dentro del repo

    # 3) commit (si hay cambios)
    # Nota: git commit falla si "no hay nada que commitear", lo manejamos:
    r = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=REPO)
    if r.returncode == 0:
        print("ℹ️ No hay cambios en stage para commitear.")
    else:
        run(["git", "commit", "-m", MENSAJE])
        run(["git", "push"])
        print("✅ add/commit/push completado.")
